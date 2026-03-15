#!/usr/bin/env python3
"""
Import Me.xlsx → Redis  (IPAM SIW v2)
======================================
Usage :
  python3 import_redis.py [--xlsx PATH] [--host 127.0.0.1] [--port 6379]
                          [--password PWD] [--dry-run] [--site NOM]

Dépendances :
  pip install openpyxl redis

Lancer depuis le serveur (ou depuis une machine qui a accès à Redis) :
  python3 import_redis.py --xlsx /tmp/Me.xlsx
  python3 import_redis.py --dry-run          # vérifier sans écrire
"""

import sys
import re
import argparse
from datetime import datetime, timezone

try:
    import openpyxl
except ImportError:
    sys.exit("Dépendance manquante : pip install openpyxl")

try:
    import redis as _redis_lib
except ImportError:
    sys.exit("Dépendance manquante : pip install redis")

# ---------------------------------------------------------------------------
# Chemin par défaut du fichier Excel  (adapter si besoin)
# ---------------------------------------------------------------------------
DEFAULT_XLSX = "Me.xlsx"
NOW = datetime.now(timezone.utc).isoformat()


# ===========================================================================
# Helpers
# ===========================================================================

def _vlan_id_from_header(cell_value):
    """Extrait l'ID VLAN numérique depuis 'VLAN 202', 'PST-03139-Vlan-1460', '202'…"""
    if cell_value is None:
        return None
    s = str(cell_value).strip()
    m = re.search(r'(?:vlan|vl)[_\s-]*(\d{1,4})\b', s, re.IGNORECASE)
    if m:
        return m.group(1)
    if re.fullmatch(r'\d{1,4}', s):
        return s
    return None


def _normalize_status(etat_cell, hostname_cell):
    """Normalise le statut IP."""
    etat = str(etat_cell).strip() if etat_cell else ''
    host = str(hostname_cell).strip().lower() if hostname_cell else ''
    if 'réserv' in etat.lower() or 'reserv' in etat.lower():
        return 'Réservée'
    if etat == 'Libre':
        return 'Libre'
    if etat == 'Utilisé':
        if 'réserv' in host or 'reserv' in host:
            return 'Réservée'
        return 'Utilisé'
    return 'Libre'


def _is_ip(val):
    """Retourne True si val ressemble à une adresse IPv4."""
    if not val:
        return False
    return bool(re.match(r'^\d{1,3}(\.\d{1,3}){3}$', str(val).strip()))


# ===========================================================================
# Parsing d'une feuille Excel
# ===========================================================================

def parse_sheet(ws):
    """
    Analyse une feuille et retourne un dict :
      { vlan_id_str: { 'network': str, 'mask': str, 'gateway': str,
                       'ips': [{'ip', 'hostname', 'status'}] } }
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    header = rows[0]
    ncols  = len(header)

    # ── 1. Détecter les blocs VLAN dans la ligne d'en-tête ────────────────
    # Un bloc commence en colonne c si :
    #   header[c+2] contient 'Etat'  ET  header[c+3] contient un VLAN ID
    block_starts = []          # [(col_index, vlan_id_str)]
    for c in range(ncols - 3):
        h2 = header[c + 2]
        h3 = header[c + 3]
        if isinstance(h2, str) and 'etat' in h2.lower():
            vid = _vlan_id_from_header(h3)
            if vid:
                block_starts.append((c, vid))

    if not block_starts:
        return {}

    vlans = {}

    # ── 2. Extraire les IPs de chaque bloc ────────────────────────────────
    for (col, vlan_id) in block_starts:
        ips = []
        for row in rows[1:]:
            if len(row) <= col:
                continue
            ip_val = row[col]
            if not _is_ip(ip_val):
                continue
            ip_str   = str(ip_val).strip()
            hostname = row[col + 1] if len(row) > col + 1 else None
            etat     = row[col + 2] if len(row) > col + 2 else None
            status   = _normalize_status(etat, hostname)
            hostname_str = str(hostname).strip() if hostname else ''
            ips.append({'ip': ip_str, 'hostname': hostname_str, 'status': status})

        if not ips:
            continue

        if vlan_id not in vlans:
            vlans[vlan_id] = {'network': '', 'mask': '', 'gateway': '', 'ips': []}
        vlans[vlan_id]['ips'].extend(ips)

    if not vlans:
        return {}

    # ── 3. Extraire la table de plages (ranges) ───────────────────────────
    # Chercher une cellule contenant 'ID VLAN' ou 'VLAN' dans la zone à droite
    # des blocs IP. On cherche sur toutes les lignes.
    last_block_col = max(c for c, _ in block_starts)
    range_search_from = last_block_col + 4   # commence après le dernier bloc

    for ri, row in enumerate(rows):
        if len(row) < range_search_from:
            continue
        for ci in range(range_search_from, len(row)):
            cell = row[ci]
            if not isinstance(cell, str):
                continue
            if not ('vlan' in cell.lower() or 'id' in cell.lower()):
                continue

            # Trouver les colonnes IP, ID VLAN, Masque, Gateway dans cette ligne
            ip_col = gw_col = mask_col = vlan_col = None
            for j, h in enumerate(row):
                if h is None:
                    continue
                hh = str(h).strip().upper()
                if re.match(r'^IP$|^RÉSEAU$|^RESEAU$|^NETWORK$', hh):
                    ip_col = j
                elif re.search(r'VLAN|VL$', hh):
                    vlan_col = j
                elif re.search(r'MASK|MASQUE', hh):
                    mask_col = j
                elif re.search(r'GATEWAY|GW|PASSERELLE', hh):
                    gw_col = j

            if vlan_col is None:
                continue
            # Fallback: colonne IP = colonne juste avant vlan_col
            if ip_col is None and vlan_col > 0:
                ip_col = vlan_col - 1

            # Lire les lignes de données de la table
            for data_row in rows[ri + 1: ri + 25]:
                if len(data_row) <= vlan_col:
                    continue
                v_raw = data_row[vlan_col]
                vid = _vlan_id_from_header(v_raw)
                if not vid:
                    continue

                net_raw  = data_row[ip_col]    if ip_col   is not None and len(data_row) > ip_col   else None
                mask_raw = data_row[mask_col]  if mask_col is not None and len(data_row) > mask_col else None
                gw_raw   = data_row[gw_col]    if gw_col   is not None and len(data_row) > gw_col   else None

                net_str  = str(net_raw).strip()  if net_raw  else ''
                mask_str = str(mask_raw).strip() if mask_raw else ''
                gw_str   = str(gw_raw).strip()   if gw_raw   else ''

                # Normaliser le CIDR
                if '/' in net_str:
                    # déjà au format CIDR, ex: "192.168.1.0/24"
                    pass
                elif mask_str.startswith('/'):
                    # masque au format "/24" → concaténer
                    net_str = net_str + mask_str

                if vid in vlans:
                    vlans[vid]['network'] = net_str
                    vlans[vid]['mask']    = mask_str
                    vlans[vid]['gateway'] = gw_str

            break  # on a trouvé la table, inutile de continuer
        else:
            continue
        break

    return vlans


# ===========================================================================
# Import Redis
# ===========================================================================

def import_site(r, site_name, vlans_data, dry_run=False):
    """Insère un site complet dans Redis. Retourne le nombre d'IPs insérées."""

    # Vérifier si le site existe déjà
    existing = r.hget('sites:idx:name', site_name.upper()) if not dry_run else None
    if existing:
        print(f"  ⚠  Site déjà présent (id={existing}), ignoré")
        return 0

    total_ips = sum(len(v['ips']) for v in vlans_data.values())

    if dry_run:
        print(f"  [DRY-RUN]  {len(vlans_data)} VLANs  |  {total_ips} IPs")
        for vid, vd in vlans_data.items():
            net = vd.get('network') or '—'
            gw  = vd.get('gateway') or '—'
            print(f"    VLAN {vid:>5}  réseau={net:<22}  gw={gw:<16}  {len(vd['ips'])} IPs")
        return total_ips

    # Créer le site
    site_id = r.incr('seq:sites')
    pipe = r.pipeline()
    pipe.hset(f'site:{site_id}', mapping={'name': site_name, 'created_at': NOW})
    pipe.hset('sites:idx:name', site_name.upper(), str(site_id))
    pipe.sadd('sites', str(site_id))
    pipe.execute()

    inserted = 0
    for vlan_id_str, vdata in vlans_data.items():
        ips = vdata['ips']
        if not ips:
            continue

        # Créer le VLAN
        vlan_db_id = r.incr('seq:vlans')
        pipe = r.pipeline()
        pipe.hset(f'vlan:{vlan_db_id}', mapping={
            'site_id':    str(site_id),
            'vlan_id':    vlan_id_str,
            'network':    vdata.get('network', ''),
            'mask':       vdata.get('mask', ''),
            'gateway':    vdata.get('gateway', ''),
            'created_at': NOW,
        })
        pipe.hset(f'site:{site_id}:vlans:idx', vlan_id_str, str(vlan_db_id))
        pipe.sadd(f'site:{site_id}:vlans', str(vlan_db_id))
        pipe.execute()

        # Insérer les IPs (dédoublonnage par adresse)
        seen_ips = set()
        for entry in ips:
            ip_addr = entry['ip']
            if ip_addr in seen_ips:
                continue
            seen_ips.add(ip_addr)

            # INSERT OR IGNORE via HSETNX
            is_new = r.hsetnx(f'vlan:{vlan_db_id}:ips:idx', ip_addr, '0')
            if not is_new:
                continue

            ip_id = r.incr('seq:ips')
            pipe = r.pipeline()
            pipe.hset(f'ip:{ip_id}', mapping={
                'vlan_id':    str(vlan_db_id),
                'ip_address': ip_addr,
                'hostname':   entry.get('hostname', ''),
                'status':     entry['status'],
                'created_at': NOW,
                'updated_at': NOW,
            })
            pipe.sadd(f'vlan:{vlan_db_id}:ips', str(ip_id))
            pipe.hset(f'vlan:{vlan_db_id}:ips:idx', ip_addr, str(ip_id))
            pipe.execute()
            inserted += 1

    return inserted


# ===========================================================================
# Main
# ===========================================================================

def main():
    parser = argparse.ArgumentParser(
        description='Importe Me.xlsx dans Redis pour IPAM SIW v2'
    )
    parser.add_argument('--xlsx',     default=DEFAULT_XLSX, help='Chemin du fichier Excel')
    parser.add_argument('--host',     default='127.0.0.1',  help='Redis host (défaut: 127.0.0.1)')
    parser.add_argument('--port',     type=int, default=6379, help='Redis port (défaut: 6379)')
    parser.add_argument('--password', default=None,          help='Mot de passe Redis (si configuré)')
    parser.add_argument('--dry-run',  action='store_true',   help='Parser sans écrire dans Redis')
    parser.add_argument('--site',     default=None,          help='Importer seulement ce site (nom de feuille, partiel)')
    args = parser.parse_args()

    # Ouvrir le fichier Excel
    print(f"\n📂 Ouverture de {args.xlsx} …")
    try:
        wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    except FileNotFoundError:
        sys.exit(f"Fichier introuvable : {args.xlsx}")

    print(f"   {len(wb.sheetnames)} feuilles : {', '.join(wb.sheetnames)}\n")

    # Connexion Redis
    if not args.dry_run:
        try:
            r = _redis_lib.Redis(
                host=args.host, port=args.port,
                password=args.password,
                decode_responses=True,
            )
            r.ping()
            print(f"✅ Redis connecté : {args.host}:{args.port}\n")
        except Exception as e:
            sys.exit(f"Impossible de se connecter à Redis : {e}")
    else:
        r = None
        print("🔍 Mode DRY-RUN — aucune écriture Redis\n")

    grand_vlans = 0
    grand_ips   = 0

    for sheet_name in wb.sheetnames:
        if args.site and args.site.lower() not in sheet_name.lower():
            continue

        site_name = sheet_name.strip().upper()
        print(f"── {site_name}")

        ws = wb[sheet_name]
        vlans_data = parse_sheet(ws)

        if not vlans_data:
            print("   (aucune donnée IP)\n")
            continue

        n = import_site(r, site_name, vlans_data, dry_run=args.dry_run)

        if not args.dry_run:
            n_vlans = len(vlans_data)
            grand_vlans += n_vlans
            grand_ips   += n
            for vid, vd in vlans_data.items():
                net = vd.get('network') or '—'
                print(f"   VLAN {vid:>5}  {net:<22}  {len(vd['ips'])} IPs")
            print(f"   → {n_vlans} VLANs, {n} IPs insérées\n")
        else:
            grand_ips += n
            grand_vlans += len(vlans_data)
            print()

    print("=" * 55)
    print(f"TOTAL : {grand_vlans} VLANs  |  {grand_ips} IPs")
    if args.dry_run:
        print("(DRY-RUN — rien n'a été écrit)")
    print()


if __name__ == '__main__':
    main()
